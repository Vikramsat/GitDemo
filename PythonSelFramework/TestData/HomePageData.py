import openpyxl


class HomePageData:
    test_HomePage_Data = [{"Firstname": "Vikram", "Email":"vik@gmail", "Password":"pwd", "Gender":"Male"},
                          {"Firstname": "Vinayak", "Email":"vin@gmail", "Password":"pwd1", "Gender":"Male"}]

    @staticmethod
    def testData(testcase_name):
        book = openpyxl.load_workbook("E:\\Udemy\\PythonDemo.xlsx")
        sheet = book.active

        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    # print(sheet.cell(row=i, column=j).value)
        print(Dict)
        return [Dict]   #we have to return it in the form of list only
